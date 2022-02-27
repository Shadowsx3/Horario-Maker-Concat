import json
from copy import deepcopy

import openpyxl
from openpyxl.styles import Alignment, Color, PatternFill, Border, Side

wb = openpyxl.Workbook()
hoja = wb.active
with open('horarios.json', encoding='utf-8') as fh:
    data = json.load(fh)
with open('colores.json', encoding='utf-8') as fh:
    colores = json.load(fh)
horas = {}
dias = {"lunes": "", "martes": "", "miercoles": "", "jueves": "", "viernes": ""}
dias2 = {"lunes": 0, "martes": 0, "miercoles": 0, "jueves": 0, "viernes": 0}
listaDia = list(dias.keys())
grupos = {}

for lugark, lugar in data.items():
    for diak, dia in lugar.items():
        for hora in dia:
            grupos[hora["nombre"]] = lugark
            if hora["inicio"] not in horas.keys():
                semana = deepcopy(dias)
                semana[diak] = hora["nombre"]
                horas[hora["inicio"]] = semana
            else:
                if horas[hora["inicio"]][diak] != "":
                    semana = deepcopy(dias)
                    semana[diak] = hora["nombre"]
                    horas[hora["inicio"] + ":00"] = semana
                else:
                    horas[hora["inicio"]][diak] = hora["nombre"]
            if hora["final"] not in horas.keys():
                semana = deepcopy(dias)
                semana[diak] = hora["nombre"]
                horas[hora["final"]] = semana
            else:
                if horas[hora["final"]][diak] != "":
                    semana = deepcopy(dias)
                    semana[diak] = hora["nombre"]
                    horas[hora["final"] + ":00"] = semana
                else:
                    horas[hora["final"]][diak] = hora["nombre"]
horasS = list(horas.keys())
horasS.sort()
# Seteo
hoja.append(('Hora', 'Lunes', 'Martes', 'Miercoles', 'Jueves', 'Viernes'))
print(horas)
actual = deepcopy(dias)
actualI = deepcopy(dias2)
merges = {}
for index, i in enumerate(horasS):
    for j in actual.keys():
        if actual[j] != "":
            pos = listaDia.index(j) + 2
            merges[f"{actualI[j]}-{pos}"] = {"start_row": actualI[j], "start_column": pos, "end_row": index + 2,
                                             "end_column": pos, "value": actual[j]}
            if horas[i][j] == actual[j]:
                horas[i][j] = ""
                actual[j] = ""
        else:
            actualI[j] = index + 2
            actual[j] = horas[i][j]
    hoja.append(
        (":".join(i.split(":")[:2]), horas[i]['lunes'], horas[i]['martes'], horas[i]['miercoles'], horas[i]['jueves'],
         horas[i]['viernes']))
    hoja.cell(index + 2, 1).border = Border(
        top=Side(style='thin'),
        bottom=Side(style='thin'))
    for m in horas[i].keys():
        cell = hoja.cell(index + 2, listaDia.index(m) + 2)
        if horas[i][m] != "":
            cell.fill = PatternFill(patternType='solid',
                                    fill_type='solid',
                                    fgColor=Color(colores[grupos[horas[i][m]]]))
            cell.border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))
        else:
            if listaDia.index(m) == len(listaDia) - 1:
                cell.border = Border(right=Side(style='thin'),
                                     top=Side(style='thin'),
                                     bottom=Side(style='thin'))
            elif listaDia.index(m) == 0:
                cell.border = Border(left=Side(style='thin'),
                                     top=Side(style='thin'),
                                     bottom=Side(style='thin'))
            else:
                cell.border = Border(
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))
for i in merges.values():
    hoja.merge_cells(start_row=i["start_row"], start_column=i["start_column"], end_row=i["end_row"],
                     end_column=i["end_column"])
    cell = hoja.cell(row=i["start_row"], column=i["start_column"], value=i["value"])
    cell.alignment = Alignment(horizontal='center', vertical='center')
wb.save('horarios.xlsx')
